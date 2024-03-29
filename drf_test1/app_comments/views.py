import csv
from datetime import datetime
from typing import List

import openpyxl
from openpyxl import Workbook

from django.http import HttpResponse
from django.views import View
from django.utils import timezone
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated, IsAuthenticatedOrReadOnly

from .models import Car, Comment, Country, Manufacturer
from .permissions import IsTokenAuthenticatedOrReadOnly
from .serializers import CarSerializer, CommentSerializer, CountrySerializer, ManufacturerSerializer


class CountryViewSet(viewsets.ModelViewSet):
    """API endpoint that allows countries to be viewed or edited."""
    queryset = Country.objects.all()
    serializer_class = CountrySerializer
    permission_classes = [IsTokenAuthenticatedOrReadOnly]


class ManufacturerViewSet(viewsets.ModelViewSet):
    """API endpoint that allows manufacturers to be viewed or edited."""
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer
    permission_classes = [IsTokenAuthenticatedOrReadOnly]


class CarViewSet(viewsets.ModelViewSet):
    """API endpoint that allows cars to be viewed or edited."""
    queryset = Car.objects.all()
    serializer_class = CarSerializer
    permission_classes = [IsTokenAuthenticatedOrReadOnly]


class CommentViewSet(viewsets.ModelViewSet):
    """API endpoint that allows comments to be viewed or edited."""
    queryset = Comment.objects.all()
    serializer_class = CommentSerializer

    def get_permissions(self):
        """Return the permissions that this view requires."""
        permission_classes = [IsAuthenticatedOrReadOnly]

        if self.action in ['update', 'partial_update', 'destroy']:
            permission_classes = [IsAuthenticated]

        return [permission() for permission in permission_classes]
    

class CarExportView(View):
    """A view for exporting cars data in CSV or XLSX format."""
    def get(self, request):
        """Handle GET requests to the view."""
        # Get the format from the query parameters
        export_format = request.GET.get('format', 'csv')

        # Get the data to be exported
        cars = Car.objects.all()

        if export_format == 'csv':
            # Export as CSV
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="cars.csv"'

            writer = csv.writer(response)
            writer.writerow(['Name', 'Manufacturer', 'Start Year', 'End Year', 'Comment Count'])

            for car in cars:
                writer.writerow([car.name, car.manufacturer.name, car.start_year, car.end_year, car.comment_set.count()])

        elif export_format == 'xlsx':
            # Export as XLSX
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="cars.xlsx"'

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'Cars'

            worksheet['A1'] = 'Name'
            worksheet['B1'] = 'Manufacturer'
            worksheet['C1'] = 'Start Year'
            worksheet['D1'] = 'End Year'
            worksheet['E1'] = 'Comment Count'

            for car in cars:
                row = [
                    car.name,
                    car.manufacturer.name,
                    car.start_year,
                    car.end_year,
                    car.comment_set.count(),
                ]
                worksheet.append(row)

            workbook.save(response)

        else:
            # Invalid format
            response = HttpResponse('Invalid export format.', content_type='text/plain', status=400)

        return response


class CountryExportView(View):
    """A view for exporting country data in CSV or XLSX format."""
    def get(self, request, *args, **kwargs):
        format = request.GET.get('format')
        if format not in ['csv', 'xlsx']:
            return HttpResponse('Unsupported format')

        # Get the data
        queryset = Country.objects.all()
        serializer = CountrySerializer(queryset, many=True)
        data = serializer.data

        # Export to CSV
        if format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="countries_{}.csv"'.format(timezone.now().strftime('%Y%m%d_%H%M%S'))
            writer = csv.writer(response)
            writer.writerow(['id', 'name'])
            for row in data:
                writer.writerow([row['id'], row['name']])
            return response

        # Export to XLSX
        if format == 'xlsx':
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="countries_{}.xlsx"'.format(timezone.now().strftime('%Y%m%d_%H%M%S'))
            wb = Workbook()
            ws = wb.active
            ws.title = 'Countries'
            ws.append(['id', 'name'])
            for row in data:
                ws.append([row['id'], row['name']])
            wb.save(response)
            return response


class ManufacturerExportView(View):
    """A view for exporting manufacturer data in CSV or XLSX format."""
    def get(self, request, *args, **kwargs):
        format = request.GET.get('format', 'csv')

        manufacturers = Manufacturer.objects.all()

        if format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="manufacturers.csv"'

            writer = csv.writer(response)
            writer.writerow(['ID', 'Name', 'Country'])

            for manufacturer in manufacturers:
                writer.writerow([manufacturer.id, manufacturer.name, manufacturer.country.name])

            return response

        elif format == 'xlsx':
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="manufacturers.xlsx"'

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Manufacturers'

            headers = ['ID', 'Name', 'Country']
            for col_num, column_title in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = column_title

            for row_num, manufacturer in enumerate(manufacturers, 1):
                row = [manufacturer.id, manufacturer.name, manufacturer.country.name]
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num+1, column=col_num)
                    cell.value = cell_value

            workbook.save(response)
            return response

        else:
            return HttpResponse(status=400, content='Bad request')


class CommentsExportView(View):
    """A view for exporting comments data in CSV or XLSX format."""
    def get(self, request, *args, **kwargs):
        format = request.GET.get('format', 'csv')
        if format not in ['csv', 'xlsx']:
            format = 'csv'
        comments = Comment.objects.all()

        if format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="comments.csv"'
            writer = csv.writer(response)
            writer.writerow(['id', 'email', 'car', 'comment', 'created_at'])
            for comment in comments:
                writer.writerow([comment.id, comment.email, comment.car.name, comment.comment, comment.created_at])
            return response

        elif format == 'xlsx':
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="comments.xlsx"'
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Comments'
            worksheet.append(['id', 'email', 'car', 'comment', 'created_at'])
            for comment in comments:
                created_at = comment.created_at.astimezone().replace(tzinfo=None)
                worksheet.append([comment.id, comment.email, comment.car.name, comment.comment, created_at])
            workbook.save(response)
            return response
